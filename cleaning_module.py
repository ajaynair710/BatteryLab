"""
Advanced Battery Data Cleaning Module for BatteryLab
Implements comprehensive 3-phase cleaning pipeline for cycler data
"""

import pandas as pd
import numpy as np
import streamlit as st
from io import BytesIO
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Union
import os
import json

# Optional imports for additional file formats
try:
    import h5py
    H5PY_AVAILABLE = True
except ImportError:
    H5PY_AVAILABLE = False

try:
    from sqlalchemy import create_engine, inspect
    SQLALCHEMY_AVAILABLE = True
except ImportError:
    SQLALCHEMY_AVAILABLE = False

try:
    import sqlite3
    SQLITE_AVAILABLE = True
except ImportError:
    SQLITE_AVAILABLE = False

# Optional imports with fallbacks
try:
    from rapidfuzz import fuzz, process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False
    def fuzz_ratio(s1, s2):
        s1_l, s2_l = str(s1).lower(), str(s2).lower()
        if s1_l == s2_l:
            return 100.0
        if s1_l in s2_l or s2_l in s1_l:
            return 80.0
        return 0.0
    
    def process_extractOne(query, choices, scorer=None):
        best_match = None
        best_score = 0.0
        for choice in choices:
            score = fuzz_ratio(query, choice)
            if score > best_score:
                best_score = score
                best_match = choice
        return (best_match, best_score) if best_match else (None, 0.0)
    
    class MockFuzz:
        ratio = staticmethod(fuzz_ratio)
    fuzz = MockFuzz()
    process = type('MockProcess', (), {'extractOne': staticmethod(process_extractOne)})()

try:
    from scipy.signal import savgol_filter
    from scipy.interpolate import interp1d
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False

try:
    from sklearn.neighbors import LocalOutlierFactor
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False


# ============================================================================
# PHASE 1: Data Ingestion and Harmonization
# ============================================================================

STANDARD_SCHEMA = {
    'DateTime': 'datetime',
    'Cycle_Index': 'int',
    'Step_Index': 'int',
    'Current': 'float',
    'Voltage': 'float',
    'Capacity': 'float',
    'Temperature': 'float',
    'Time': 'float'
}

# Comprehensive Column Alias Mapping
# Maps various column name variations to standard names
COLUMN_ALIAS_MAP = {
    # Voltage aliases
    'Voltage': ['voltage', 'v', 'volt', 'volts', 'u', 'u_cell', 'u_cell_v', 'cell_voltage', 
                'potential', 'potential_v', 'e', 'ewe', 'ewe_v', 'v_cell', 'vcell', 
                'voltage_v', 'voltage(v)', 'v_measured', 'measured_voltage', 'battery_voltage',
                'terminal_voltage', 'ocv', 'open_circuit_voltage', 'v_terminal', 'v_battery'],
    
    # Current aliases
    'Current': ['current', 'i', 'i_measured', 'current_a', 'current(a)', 'i_a', 'i_ma', 
                'current_ma', 'amps', 'ampere', 'amperes', 'i_cell', 'icell', 'i_charge',
                'i_discharge', 'current_charge', 'current_discharge', 'i_load', 'load_current'],
    
    # Capacity aliases
    'Capacity': ['capacity', 'q', 'q_charge', 'q_discharge', 'capacity_ah', 'capacity(ah)',
                 'q_ah', 'charge_capacity', 'discharge_capacity', 'capa', 'cap', 'q_cell',
                 'qcell', 'capacity_charge', 'capacity_discharge', 'q_charge_ah', 'q_discharge_ah',
                 'cumulative_capacity', 'total_capacity', 'q_total'],
    
    # Temperature aliases
    'Temperature': ['temperature', 'temp', 't', 't_cell', 'tcell', 'temp_c', 'temperature_c',
                    'temp_celsius', 'temperature_celsius', 't_measured', 'measured_temp',
                    'cell_temperature', 'battery_temperature', 't_ambient', 'ambient_temp',
                    't_surface', 'surface_temp', 't_internal', 'internal_temp'],
    
    # Time aliases
    'Time': ['time', 't', 'test_time', 'test_time_s', 'time_s', 'time(s)', 'elapsed_time',
             'elapsed_time_s', 'step_time', 'step_time_s', 'step_time(s)', 'relative_time',
             'time_elapsed', 'duration', 't_elapsed', 'time_relative', 't_relative'],
    
    # DateTime aliases
    'DateTime': ['datetime', 'date_time', 'timestamp', 'time_stamp', 'date', 'time_full',
                 'test_date', 'measurement_time', 'record_time', 'log_time', 'acquisition_time',
                 'data_time', 'sample_time', 't_datetime', 'dt', 'date_time_full'],
    
    # Cycle aliases
    'Cycle_Index': ['cycle', 'cycle_index', 'cycle_number', 'cycle_num', 'n_cycle', 'ncycle',
                    'cycle_id', 'cycleid', 'cycle_idx', 'cycleindex', 'cycle_number', 'n',
                    'cycle_count', 'cycle_count', 'cycle_no', 'cycleno', 'c_cycle'],
    
    # Step aliases
    'Step_Index': ['step', 'step_index', 'step_number', 'step_num', 'n_step', 'nstep',
                   'step_id', 'stepid', 'step_idx', 'stepindex', 'step_number', 'ns',
                   'step_count', 'step_no', 'stepno', 's_step', 'test_step', 'step_type_id'],
}

# Create reverse lookup for quick access
ALIAS_TO_STANDARD = {}
for standard_name, aliases in COLUMN_ALIAS_MAP.items():
    for alias in aliases:
        ALIAS_TO_STANDARD[alias.lower()] = standard_name


def read_file_universal(file_path_or_buffer, file_type: str = None, **kwargs) -> pd.DataFrame:
    """
    Universal file reader supporting multiple formats.
    
    Supports:
    - Excel (.xlsx, .xls)
    - CSV (.csv, .tsv)
    - JSON (.json)
    - Text files (.txt, .dat)
    - HDF5 (.h5, .hdf5)
    - SQL databases (SQLite, PostgreSQL, MySQL)
    - MATLAB (.mat)
    
    Parameters:
    -----------
    file_path_or_buffer : str, path-like, or file-like object
        Path to file or file buffer
    file_type : str, optional
        File type hint ('excel', 'csv', 'json', 'txt', 'hdf5', 'sql', 'mat').
        If None, auto-detects from extension.
    **kwargs : dict
        Additional arguments passed to specific readers
    
    Returns:
    --------
    pd.DataFrame
        Loaded dataframe
    """
    # Auto-detect file type if not provided
    if file_type is None:
        if isinstance(file_path_or_buffer, str):
            ext = os.path.splitext(file_path_or_buffer)[1].lower()
        elif hasattr(file_path_or_buffer, 'name'):
            ext = os.path.splitext(file_path_or_buffer.name)[1].lower()
        else:
            raise ValueError("Cannot determine file type. Please specify file_type parameter.")
        
        ext_to_type = {
            '.xlsx': 'excel', '.xls': 'excel',
            '.csv': 'csv', '.tsv': 'csv',
            '.json': 'json',
            '.txt': 'txt', '.dat': 'txt',
            '.h5': 'hdf5', '.hdf5': 'hdf5',
            '.db': 'sql', '.sqlite': 'sql', '.sqlite3': 'sql',
            '.mat': 'mat'
        }
        file_type = ext_to_type.get(ext, 'csv')  # Default to CSV
    
    file_type = file_type.lower()
    
    # Read based on file type
    if file_type == 'excel':
        sheet_name = kwargs.get('sheet_name', 0)
        header = kwargs.get('header', 0)
        return pd.read_excel(file_path_or_buffer, sheet_name=sheet_name, header=header)
    
    elif file_type == 'csv':
        sep = kwargs.get('sep', ',')
        header = kwargs.get('header', 0)
        return pd.read_csv(file_path_or_buffer, sep=sep, header=header)
    
    elif file_type == 'json':
        orient = kwargs.get('orient', 'records')
        return pd.read_json(file_path_or_buffer, orient=orient)
    
    elif file_type == 'txt':
        # Try to detect delimiter
        sep = kwargs.get('sep', None)
        if sep is None:
            # Read first line to detect delimiter
            if isinstance(file_path_or_buffer, str):
                with open(file_path_or_buffer, 'r') as f:
                    first_line = f.readline()
            else:
                pos = file_path_or_buffer.tell()
                first_line = file_path_or_buffer.readline()
                file_path_or_buffer.seek(pos)
            
            # Common delimiters
            for delimiter in [',', '\t', ';', '|', ' ']:
                if delimiter in first_line:
                    sep = delimiter
                    break
            if sep is None:
                sep = '\s+'  # Whitespace
        
        header = kwargs.get('header', 0)
        return pd.read_csv(file_path_or_buffer, sep=sep, header=header, engine='python')
    
    elif file_type == 'hdf5':
        if not H5PY_AVAILABLE:
            raise ImportError("h5py is required for HDF5 files. Install with: pip install h5py")
        
        key = kwargs.get('key', None)
        if isinstance(file_path_or_buffer, str):
            with h5py.File(file_path_or_buffer, 'r') as f:
                if key is None:
                    # List available keys
                    keys = list(f.keys())
                    if len(keys) == 0:
                        raise ValueError("No datasets found in HDF5 file")
                    key = keys[0]  # Use first key
                
                data = f[key][:]
                # Convert to DataFrame
                if isinstance(data, np.ndarray):
                    if data.dtype.names:  # Structured array
                        return pd.DataFrame(data)
                    else:
                        # Try to infer column names
                        return pd.DataFrame(data, columns=[f'col_{i}' for i in range(data.shape[1])])
                else:
                    return pd.DataFrame(data)
        else:
            raise ValueError("HDF5 files require a file path, not a buffer")
    
    elif file_type == 'sql':
        if not SQLALCHEMY_AVAILABLE and not SQLITE_AVAILABLE:
            raise ImportError("sqlalchemy or sqlite3 is required for SQL files")
        
        table_name = kwargs.get('table_name', None)
        query = kwargs.get('query', None)
        
        if isinstance(file_path_or_buffer, str):
            # Detect SQLite vs other databases
            if file_path_or_buffer.endswith(('.db', '.sqlite', '.sqlite3')):
                if SQLITE_AVAILABLE:
                    conn = sqlite3.connect(file_path_or_buffer)
                    if query:
                        return pd.read_sql_query(query, conn)
                    elif table_name:
                        return pd.read_sql_table(table_name, conn)
                    else:
                        # List tables and use first one
                        cursor = conn.cursor()
                        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                        tables = cursor.fetchall()
                        if len(tables) == 0:
                            raise ValueError("No tables found in SQLite database")
                        return pd.read_sql_table(tables[0][0], conn)
                else:
                    raise ImportError("sqlite3 is required for SQLite files")
            else:
                # Other SQL databases (PostgreSQL, MySQL, etc.)
                if not SQLALCHEMY_AVAILABLE:
                    raise ImportError("sqlalchemy is required for SQL databases")
                
                connection_string = kwargs.get('connection_string', None)
                if connection_string is None:
                    raise ValueError("connection_string required for non-SQLite databases")
                
                engine = create_engine(connection_string)
                if query:
                    return pd.read_sql_query(query, engine)
                elif table_name:
                    return pd.read_sql_table(table_name, engine)
                else:
                    # List tables and use first one
                    inspector = inspect(engine)
                    tables = inspector.get_table_names()
                    if len(tables) == 0:
                        raise ValueError("No tables found in database")
                    return pd.read_sql_table(tables[0], engine)
        else:
            raise ValueError("SQL databases require a connection string or file path")
    
    elif file_type == 'mat':
        try:
            from scipy.io import loadmat
        except ImportError:
            raise ImportError("scipy is required for MATLAB files. Install with: pip install scipy")
        
        data = loadmat(file_path_or_buffer)
        # Remove MATLAB metadata keys
        keys = [k for k in data.keys() if not k.startswith('__')]
        if len(keys) == 0:
            raise ValueError("No data found in MATLAB file")
        
        # Use first data key
        mat_data = data[keys[0]]
        if isinstance(mat_data, np.ndarray):
            if mat_data.ndim == 2:
                return pd.DataFrame(mat_data)
            else:
                raise ValueError(f"MATLAB data has {mat_data.ndim} dimensions. Expected 2D array.")
        else:
            raise ValueError("Unexpected MATLAB data format")
    
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


CYCLER_PARSERS = {
    'calce': {
        'cycle_cols': ['Cycle', 'Cycle_Index', 'Cycle Number'],
        'step_cols': ['Step', 'Step_Index', 'Step Number'],
        'current_cols': ['Current', 'I', 'Current(A)', 'Current_A'],
        'voltage_cols': ['Voltage', 'V', 'Voltage(V)', 'Voltage_V'],
        'capacity_cols': ['Capacity', 'Q', 'Capacity(Ah)', 'Capacity_Ah'],
        'temp_cols': ['Temperature', 'Temp', 'Temperature(C)', 'Temperature_C'],
        'time_cols': ['Step_Time', 'Time', 'Test_Time', 'Time(s)'],
        'datetime_cols': ['Step_Time'],  # Excel serial time
        'excel_serial_origin': '1899-12-30'  # CALCE uses Excel serial time
    },
    'arbin': {
        'cycle_cols': ['Cycle_Index', 'Cycle', 'Cycle_Number'],
        'step_cols': ['Step_Index', 'Step', 'Step_Number'],
        'current_cols': ['Current', 'I', 'Current(A)', 'Current_A'],
        'voltage_cols': ['Voltage', 'V', 'Voltage(V)', 'Voltage_V'],
        'capacity_cols': ['Capacity', 'Q', 'Capacity(Ah)', 'Capacity_Ah'],
        'temp_cols': ['Temperature', 'Temp', 'Temperature(C)', 'Temperature_C'],
        'time_cols': ['Test_Time', 'Time', 'Test_Time(s)', 'Time(s)']
    },
    'biologic': {
        'cycle_cols': ['cycle number', 'Cycle', 'N'],
        'step_cols': ['step number', 'Step', 'Ns'],
        'current_cols': ['I/mA', 'I', 'Current'],
        'voltage_cols': ['Ewe/V', 'Ewe', 'Voltage'],
        'capacity_cols': ['Q charge/discharge/mAh', 'Q', 'Capacity'],
        'temp_cols': ['T', 'Temperature'],
        'time_cols': ['time/s', 'Time', 't']
    },
    'neware': {
        'cycle_cols': ['Cycle', 'Cycle_Index', 'Cycle Number'],
        'step_cols': ['Step', 'Step_Index', 'Step Number'],
        'current_cols': ['Current', 'I', 'Current(A)'],
        'voltage_cols': ['Voltage', 'V', 'Voltage(V)'],
        'capacity_cols': ['Capacity', 'Q', 'Capacity(Ah)'],
        'temp_cols': ['Temperature', 'Temp'],
        'time_cols': ['Time', 'Test_Time', 'Time(s)']
    },
    'maccor': {
        'cycle_cols': ['Cycle', 'Cyc', 'Cycle Number'],
        'step_cols': ['Step', 'Step Number'],
        'current_cols': ['Amps', 'Current', 'I'],
        'voltage_cols': ['Volts', 'Voltage', 'V'],
        'capacity_cols': ['Amp-hrs', 'Capacity', 'Q'],
        'temp_cols': ['Temp', 'Temperature'],
        'time_cols': ['Time', 'Test Time', 'Time(s)']
    },
    'solartron': {
        'cycle_cols': ['Cycle', 'Cycle Number'],
        'step_cols': ['Step', 'Step Number'],
        'current_cols': ['Current', 'I'],
        'voltage_cols': ['Voltage', 'V'],
        'capacity_cols': ['Capacity', 'Q'],
        'temp_cols': ['Temperature', 'Temp'],
        'time_cols': ['Time', 'Test Time']
    }
}


def detect_cycler_format(df: pd.DataFrame, metadata: dict = None) -> str:
    """
    Detect which cycler format the data is from.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    metadata : dict, optional
        Metadata dict that may contain format hints (e.g., from Excel headers)
    
    Returns:
    --------
    str
        Detected cycler format ('calce', 'arbin', 'biologic', 'neware', 'maccor', 'solartron', 'severson', 'unknown')
    """
    cols_l = [c.lower() for c in df.columns]
    
    # Check for CALCE format first (has Step_Time column with Excel serial time)
    if 'step_time' in cols_l:
        # Check if values look like Excel serial days (typically > 40000 for modern dates)
        step_time_col = None
        for col in df.columns:
            if col.lower() == 'step_time':
                step_time_col = col
                break
        
        if step_time_col is not None:
            sample_values = df[step_time_col].dropna().head(100)
            if len(sample_values) > 0:
                # Excel serial days are typically > 40000 for dates after 2009
                if sample_values.min() > 40000 or (sample_values.min() > 1000 and sample_values.max() < 1000000):
                    return 'calce'
    
    # Check metadata for format hints
    if metadata:
        if 'format' in metadata and metadata['format']:
            return metadata['format'].lower()
        if 'cycler' in metadata and metadata['cycler']:
            cycler_name = str(metadata['cycler']).lower()
            if 'calce' in cycler_name:
                return 'calce'
    
    # Check for Severson .mat format (has specific structure)
    if 'cycle_life' in cols_l or 'capacity' in cols_l:
        # Additional checks for Severson format
        if any('summary' in c for c in cols_l) or any('cycles' in c for c in cols_l):
            return 'severson'
    
    # Check other cycler formats
    for cycler, patterns in CYCLER_PARSERS.items():
        if cycler == 'calce':  # Already checked
            continue
        matches = 0
        for key, col_list in patterns.items():
            if key in ['excel_serial_origin', 'datetime_cols']:
                continue
            for col_pattern in col_list:
                if any(col_pattern.lower() in c for c in cols_l):
                    matches += 1
                    break
        if matches >= 4:  # At least 4 out of 7 categories match
            return cycler
    
    return 'unknown'


def harmonize_to_standard_schema(df: pd.DataFrame, cycler_format: str = None) -> pd.DataFrame:
    """
    Map cycler-specific columns to standardized schema using comprehensive alias mapping.
    
    ENHANCED v2.0: Uses comprehensive COLUMN_ALIAS_MAP for flexible column name recognition
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    cycler_format : str, optional
        Cycler format. If None, auto-detects.
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with standardized column names
    """
    if cycler_format is None:
        cycler_format = detect_cycler_format(df)
    
    df_clean = df.copy()
    cols_l = [str(c).lower() if c is not None else '' for c in df_clean.columns]
    mapping = {}
    
    # First, try comprehensive alias mapping
    for col_idx, col_name in enumerate(df_clean.columns):
        col_lower = str(col_name).lower() if col_name is not None else ''
        
        # Check if column name matches any alias
        if col_lower in ALIAS_TO_STANDARD:
            standard_name = ALIAS_TO_STANDARD[col_lower]
            # Only map if not already mapped
            if standard_name not in mapping.values():
                mapping[col_name] = standard_name
                continue
        
        # Also check for partial matches (e.g., "Voltage(V)" contains "voltage")
        for standard_name, aliases in COLUMN_ALIAS_MAP.items():
            if standard_name not in mapping.values():  # Don't remap
                for alias in aliases:
                    if alias.lower() in col_lower or col_lower in alias.lower():
                        mapping[col_name] = standard_name
                        break
                if standard_name in mapping.values():
                    break
    
    # Then, try cycler-specific patterns if format is known
    if cycler_format != 'unknown' and cycler_format in CYCLER_PARSERS:
        patterns = CYCLER_PARSERS[cycler_format]
        
        # Map cycle
        if 'Cycle_Index' not in mapping.values():
            for pattern in patterns['cycle_cols']:
                for i, col_l in enumerate(cols_l):
                    if pattern.lower() in col_l:
                        mapping[df_clean.columns[i]] = 'Cycle_Index'
                        break
                if 'Cycle_Index' in mapping.values():
                    break
        
        # Map step
        if 'Step_Index' not in mapping.values():
            for pattern in patterns['step_cols']:
                for i, col_l in enumerate(cols_l):
                    if pattern.lower() in col_l:
                        mapping[df_clean.columns[i]] = 'Step_Index'
                        break
                if 'Step_Index' in mapping.values():
                    break
        
        # Map current
        if 'Current' not in mapping.values():
            for pattern in patterns['current_cols']:
                for i, col_l in enumerate(cols_l):
                    if pattern.lower() in col_l:
                        mapping[df_clean.columns[i]] = 'Current'
                        break
                if 'Current' in mapping.values():
                    break
        
        # Map voltage
        if 'Voltage' not in mapping.values():
            for pattern in patterns['voltage_cols']:
                for i, col_l in enumerate(cols_l):
                    if pattern.lower() in col_l:
                        mapping[df_clean.columns[i]] = 'Voltage'
                        break
                if 'Voltage' in mapping.values():
                    break
        
        # Map capacity
        if 'Capacity' not in mapping.values():
            for pattern in patterns['capacity_cols']:
                for i, col_l in enumerate(cols_l):
                    if pattern.lower() in col_l or ('capacity' in col_l and 'discharge' not in col_l):
                        mapping[df_clean.columns[i]] = 'Capacity'
                        break
                if 'Capacity' in mapping.values():
                    break
        
        # Map temperature
        if 'Temperature' not in mapping.values():
            for pattern in patterns['temp_cols']:
                for i, col_l in enumerate(cols_l):
                    if pattern.lower() in col_l:
                        mapping[df_clean.columns[i]] = 'Temperature'
                        break
                if 'Temperature' in mapping.values():
                    break
        
        # Map time
        if 'Time' not in mapping.values():
            for pattern in patterns['time_cols']:
                for i, col_l in enumerate(cols_l):
                    if pattern.lower() in col_l:
                        mapping[df_clean.columns[i]] = 'Time'
                        break
                if 'Time' in mapping.values():
                    break
    
    # Apply mapping
    df_clean = df_clean.rename(columns=mapping)
    
    # Handle CALCE Excel serial time conversion (CRITICAL FIX for Issue 1)
    if cycler_format == 'calce' and 'Step_Time' in df_clean.columns:
        # CALCE uses Excel serial time (days since 1899-12-30)
        origin = pd.Timestamp("1899-12-30")
        df_clean['DateTime'] = origin + pd.to_timedelta(df_clean['Step_Time'], unit="D")
        # Also create Time column in seconds from DateTime
        if 'Time' not in df_clean.columns:
            df_clean['Time'] = (df_clean['DateTime'] - df_clean['DateTime'].min()).dt.total_seconds()
    else:
        # Normalize time to seconds for other formats
        if 'Time' in df_clean.columns:
            df_clean = normalize_time_to_seconds(df_clean)
        
        # v1.2 Fix #3: Use Date_Time if available, don't create synthetic DateTime
        if 'Date_Time' in df_clean.columns:
            # Parse Date_Time as canonical timestamp
            df_clean['DateTime'] = pd.to_datetime(df_clean['Date_Time'], errors='coerce')
        elif 'DateTime' not in df_clean.columns:
            if 'Time' in df_clean.columns:
                # Only create synthetic DateTime if no Date_Time exists
                # Use first sample's time as reference if available
                test_start = pd.Timestamp.now() - pd.Timedelta(seconds=df_clean['Time'].max())
                df_clean['DateTime'] = test_start + pd.to_timedelta(df_clean['Time'], unit='s')
            else:
                df_clean['DateTime'] = pd.Timestamp.now()
    
    # v1.2 Fix #3: If both Date_Time and DateTime exist, prefer Date_Time
    if 'Date_Time' in df_clean.columns and 'DateTime' in df_clean.columns:
        # Overwrite DateTime with parsed Date_Time
        df_clean['DateTime'] = pd.to_datetime(df_clean['Date_Time'], errors='coerce')
    
    return df_clean


def export_to_vdf(df: pd.DataFrame, file_path: str, metadata: dict = None) -> None:
    """
    Export dataframe to Voltaiq Data Format (VDF) - standardized CSV format.
    
    VDF is a standardized CSV schema for battery test data that captures variability
    across labs and equipment. This function converts cleaned data to VDF format.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Cleaned dataframe with standardized schema
    file_path : str
        Output file path (.csv)
    metadata : dict, optional
        Additional metadata to include in VDF header
    """
    # VDF standard columns (Voltaiq Data Format)
    vdf_columns = {
        'test_time': 'Time',
        'date_time': 'DateTime',
        'cycle_index': 'Cycle_Index',
        'step_index': 'Step_Index',
        'current': 'Current',
        'voltage': 'Voltage',
        'charge_capacity': 'Capacity',
        'discharge_capacity': 'Discharge_Capacity',
        'temperature': 'Temperature',
        'power': 'Power',
        'energy': 'Energy'
    }
    
    # Create VDF dataframe
    vdf_df = pd.DataFrame()
    
    # Map standard columns to VDF columns
    for vdf_col, std_col in vdf_columns.items():
        if std_col in df.columns:
            vdf_df[vdf_col] = df[std_col]
    
    # Add any additional columns
    for col in df.columns:
        if col not in vdf_columns.values():
            vdf_df[col] = df[col]
    
    # Write CSV with metadata header
    with open(file_path, 'w', encoding='utf-8') as f:
        # Write VDF header comments
        f.write("# Voltaiq Data Format (VDF) v1.0\n")
        f.write("# Standardized battery test data format\n")
        f.write(f"# Generated: {datetime.now().isoformat()}\n")
        
        if metadata:
            for key, value in metadata.items():
                f.write(f"# {key}: {value}\n")
        
        f.write("#\n")
        
        # Write data
        vdf_df.to_csv(f, index=False, lineterminator='\n')


def import_from_vdf(file_path: str) -> Tuple[pd.DataFrame, dict]:
    """
    Import dataframe from Voltaiq Data Format (VDF) - standardized CSV format.
    
    Parameters:
    -----------
    file_path : str
        Path to VDF file (.csv)
    
    Returns:
    --------
    Tuple[pd.DataFrame, dict]
        Dataframe and metadata dictionary
    """
    metadata = {}
    
    # Read header comments
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = []
        for line in f:
            if line.startswith('#'):
                # Parse metadata
                if ':' in line:
                    key, value = line[1:].strip().split(':', 1)
                    metadata[key.strip()] = value.strip()
                lines.append(line)
            else:
                break
    
    # Read CSV data
    df = pd.read_csv(file_path, comment='#')
    
    # Map VDF columns back to standard columns
    vdf_to_standard = {
        'test_time': 'Time',
        'date_time': 'DateTime',
        'cycle_index': 'Cycle_Index',
        'step_index': 'Step_Index',
        'current': 'Current',
        'voltage': 'Voltage',
        'charge_capacity': 'Capacity',
        'discharge_capacity': 'Discharge_Capacity',
        'temperature': 'Temperature',
        'power': 'Power',
        'energy': 'Energy'
    }
    
    df_renamed = df.rename(columns=vdf_to_standard)
    
    return df_renamed, metadata


def normalize_time_to_seconds(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize time column to seconds relative to test start.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with 'Time' column
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with normalized time in seconds
    """
    df_clean = df.copy()
    
    if 'Time' in df_clean.columns:
        time_col = df_clean['Time']
        
        # Detect time unit from column name or values
        time_unit = 'seconds'
        if time_col.max() > 86400:  # Likely in seconds already
            time_unit = 'seconds'
        elif time_col.max() > 1440:  # Likely in minutes
            time_unit = 'minutes'
        elif time_col.max() > 24:  # Likely in hours
            time_unit = 'hours'
        
        # Convert to seconds
        if time_unit == 'minutes':
            df_clean['Time'] = time_col * 60
        elif time_unit == 'hours':
            df_clean['Time'] = time_col * 3600
        
        # Normalize to start at 0
        df_clean['Time'] = df_clean['Time'] - df_clean['Time'].min()
    
    return df_clean


# ============================================================================
# PHASE 2: Anomaly Detection and Correction
# ============================================================================

def segment_steps_and_cycles(df: pd.DataFrame) -> pd.DataFrame:
    """
    Identify and correct step boundaries and cycle isolation.
    Creates clean cycle boundaries for proper cycle-level analysis.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with standardized schema
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with corrected step boundaries and cycle isolation
    """
    df_clean = df.copy()
    
    # Sort by time first
    if 'Time' in df_clean.columns:
        df_clean = df_clean.sort_values('Time').reset_index(drop=True)
    elif 'DateTime' in df_clean.columns:
        df_clean = df_clean.sort_values('DateTime').reset_index(drop=True)
    
    # Detect step boundaries based on current changes
    if 'Current' in df_clean.columns:
        current_diff = df_clean['Current'].diff().abs()
        step_threshold = current_diff.quantile(0.95)  # 95th percentile as threshold
        
        # Identify step boundaries
        step_boundaries = current_diff > step_threshold
        
        # Correct Step_Index if missing or incorrect
        if 'Step_Index' in df_clean.columns:
            step_idx = df_clean['Step_Index'].iloc[0] if len(df_clean) > 0 else 1
            for i in range(1, len(df_clean)):
                if step_boundaries.iloc[i]:
                    step_idx += 1
                df_clean.loc[i, 'Step_Index'] = step_idx
        else:
            step_idx = 1
            step_indices = [1]
            for i in range(1, len(df_clean)):
                if step_boundaries.iloc[i]:
                    step_idx += 1
                step_indices.append(step_idx)
            df_clean['Step_Index'] = step_indices
    
    # Correct Cycle_Index if missing or create clean cycle boundaries (Issue 2 Fix)
    if 'Cycle_Index' not in df_clean.columns or df_clean['Cycle_Index'].isna().any():
        if 'Step_Index' in df_clean.columns:
            # Assume cycle changes when step resets to 1 or when current sign changes
            cycle_idx = 1
            cycle_indices = [1]
            prev_step = df_clean['Step_Index'].iloc[0] if len(df_clean) > 0 else 1
            prev_current_sign = 0
            
            for i in range(1, len(df_clean)):
                current_step = df_clean['Step_Index'].iloc[i]
                current_sign = 1 if df_clean['Current'].iloc[i] > 0.01 else (-1 if df_clean['Current'].iloc[i] < -0.01 else 0)
                
                # New cycle if step resets OR if we transition from charge to discharge (or vice versa)
                if current_step < prev_step or (prev_current_sign != 0 and current_sign != 0 and current_sign != prev_current_sign):
                    cycle_idx += 1
                
                cycle_indices.append(cycle_idx)
                prev_step = current_step
                prev_current_sign = current_sign
            
            df_clean['Cycle_Index'] = cycle_indices
        else:
            # No step info, create cycles based on current sign changes
            if 'Current' in df_clean.columns:
                cycle_idx = 1
                cycle_indices = [1]
                prev_sign = 1 if df_clean['Current'].iloc[0] > 0.01 else (-1 if df_clean['Current'].iloc[0] < -0.01 else 0)
                
                for i in range(1, len(df_clean)):
                    current_sign = 1 if df_clean['Current'].iloc[i] > 0.01 else (-1 if df_clean['Current'].iloc[i] < -0.01 else 0)
                    if prev_sign != 0 and current_sign != 0 and current_sign != prev_sign:
                        cycle_idx += 1
                    cycle_indices.append(cycle_idx)
                    prev_sign = current_sign
                
                df_clean['Cycle_Index'] = cycle_indices
    
    # Ensure Cycle_Index is integer
    if 'Cycle_Index' in df_clean.columns:
        df_clean['Cycle_Index'] = df_clean['Cycle_Index'].astype(int)
    
    return df_clean


def assign_step_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Assign step types based on current and voltage patterns.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with 'Step_Type' column
    """
    df_clean = df.copy()
    
    if 'Current' not in df_clean.columns or 'Voltage' not in df_clean.columns:
        return df_clean
    
    step_types = []
    
    # Extract columns as numpy arrays for safe scalar access
    current_values = df_clean['Current'].values if 'Current' in df_clean.columns else np.zeros(len(df_clean))
    voltage_values = df_clean['Voltage'].values if 'Voltage' in df_clean.columns else np.zeros(len(df_clean))
    
    for idx in range(len(df_clean)):
        # Safely extract scalar values
        current_val = current_values[idx]
        voltage_val = voltage_values[idx]
        
        # Handle array/scalar cases and convert to float
        if isinstance(current_val, (np.ndarray, list)):
            current_val = float(current_val[0]) if len(current_val) > 0 else 0.0
        else:
            try:
                current_val = float(current_val) if not pd.isna(current_val) else 0.0
            except (ValueError, TypeError):
                current_val = 0.0
        
        if isinstance(voltage_val, (np.ndarray, list)):
            voltage_val = float(voltage_val[0]) if len(voltage_val) > 0 else 0.0
        else:
            try:
                voltage_val = float(voltage_val) if not pd.isna(voltage_val) else 0.0
            except (ValueError, TypeError):
                voltage_val = 0.0
        
        if abs(current_val) < 0.01:  # Near zero current
            step_type = 'Rest'
        elif current_val > 0.01:  # Positive current
            if voltage_val > 4.0:  # High voltage
                step_type = 'Constant Voltage Charge'
            else:
                step_type = 'Constant Current Charge'
        elif current_val < -0.01:  # Negative current
            if voltage_val < 2.5:  # Low voltage
                step_type = 'Constant Voltage Discharge'
            else:
                step_type = 'Constant Current Discharge'
        else:
            step_type = 'Unknown'
        
        step_types.append(step_type)
    
    df_clean['Step_Type'] = step_types
    return df_clean


def identify_rpt_cycles(df: pd.DataFrame) -> pd.DataFrame:
    """
    Identify Reference Performance Test (RPT) cycles.
    RPTs are typically longer cycles with different protocols.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with 'Is_RPT' boolean column
    """
    df_clean = df.copy()
    
    if 'Cycle_Index' not in df_clean.columns or 'Time' not in df_clean.columns:
        df_clean['Is_RPT'] = False
        return df_clean
    
    # Calculate cycle duration
    cycle_durations = df_clean.groupby('Cycle_Index')['Time'].agg(['min', 'max'])
    cycle_durations['duration'] = cycle_durations['max'] - cycle_durations['min']
    
    # RPT cycles are typically longer (outliers in duration)
    median_duration = cycle_durations['duration'].median()
    std_duration = cycle_durations['duration'].std()
    rpt_threshold = median_duration + 2 * std_duration
    
    # Mark RPT cycles
    is_rpt = df_clean['Cycle_Index'].map(
        lambda x: cycle_durations.loc[x, 'duration'] > rpt_threshold if x in cycle_durations.index else False
    )
    
    df_clean['Is_RPT'] = is_rpt
    return df_clean


def detect_statistical_anomalies(df: pd.DataFrame, window_size: int = 10, method: str = 'zscore') -> pd.DataFrame:
    """
    Detect statistical anomalies using LOF or Z-Score method.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    window_size : int
        Window size for moving window analysis
    method : str
        'zscore' or 'lof' (Local Outlier Factor)
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with 'Anomaly_Flag' column
    """
    df_clean = df.copy()
    df_clean['Anomaly_Flag'] = False
    
    numeric_cols = ['Voltage', 'Current', 'Capacity', 'Temperature']
    available_cols = [c for c in numeric_cols if c in df_clean.columns]
    
    for col in available_cols:
        # Ensure column is numeric before processing
        if col not in df_clean.columns:
            continue
        
        # Try to convert to numeric if needed
        col_data = df_clean[col]
        if not pd.api.types.is_numeric_dtype(col_data):
            # Try to convert to numeric
            col_data_numeric = pd.to_numeric(col_data, errors='coerce')
            if col_data_numeric.isna().all():
                # Can't convert to numeric, skip this column
                continue
            # Use converted numeric data
            col_data = col_data_numeric
        else:
            # Already numeric, but ensure it's a Series
            col_data = pd.Series(col_data.values, index=df_clean.index) if not isinstance(col_data, pd.Series) else col_data
        
        if method == 'zscore' and len(df_clean) > window_size:
            # Z-Score method on moving window
            try:
                rolling_mean = col_data.rolling(window=window_size, center=True).mean()
                rolling_std = col_data.rolling(window=window_size, center=True).std()
                z_scores = (col_data - rolling_mean) / (rolling_std + 1e-10)
                
                # Flag outliers (|z| > 3) - ensure boolean mask is properly formed
                anomaly_mask = abs(z_scores) > 3
                # Convert to numpy boolean array to avoid any pandas indexing issues
                if isinstance(anomaly_mask, pd.Series):
                    anomaly_mask_array = anomaly_mask.values
                else:
                    anomaly_mask_array = np.asarray(anomaly_mask, dtype=bool)
                
                # Flatten if needed
                if isinstance(anomaly_mask_array, np.ndarray) and anomaly_mask_array.ndim > 1:
                    anomaly_mask_array = anomaly_mask_array.flatten()
                
                # Ensure length matches
                if len(anomaly_mask_array) == len(df_clean):
                    df_clean.loc[anomaly_mask_array, 'Anomaly_Flag'] = True
            except Exception as e:
                # Skip this column if rolling statistics fail
                continue
        
        elif method == 'lof' and SKLEARN_AVAILABLE and len(df_clean) > 20:
            # Local Outlier Factor method
            try:
                # Ensure data is numeric
                if not pd.api.types.is_numeric_dtype(col_data):
                    col_data = pd.to_numeric(col_data, errors='coerce')
                
                # Use sliding window approach for time series
                for i in range(window_size, len(df_clean) - window_size):
                    window_data = col_data.iloc[i-window_size:i+window_size].values.reshape(-1, 1)
                    # Remove NaN values
                    window_data_clean = window_data[~np.isnan(window_data.flatten())]
                    if len(window_data_clean) > 5:
                        window_data_clean = window_data_clean.reshape(-1, 1)
                        lof = LocalOutlierFactor(n_neighbors=min(5, len(window_data_clean)-1))
                        pred = lof.fit_predict(window_data_clean)
                        if len(pred) > window_size and pred[window_size] == -1:  # Outlier detected
                            df_clean.loc[i, 'Anomaly_Flag'] = True
            except Exception:
                pass
    
    return df_clean


def correct_anomalies(df: pd.DataFrame, method: str = 'savgol') -> pd.DataFrame:
    """
    Correct detected anomalies using smoothing or interpolation.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with 'Anomaly_Flag' column
    method : str
        'savgol', 'interpolate', or 'median_filter'
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with corrected values
    """
    df_clean = df.copy()
    
    if 'Anomaly_Flag' not in df_clean.columns:
        return df_clean
    
    numeric_cols = ['Voltage', 'Current', 'Capacity', 'Temperature']
    available_cols = [c for c in numeric_cols if c in df_clean.columns]
    
    for col in available_cols:
        anomaly_mask = df_clean['Anomaly_Flag']
        
        if method == 'savgol' and SCIPY_AVAILABLE:
            # Savitzky-Golay filter for small anomalies
            if anomaly_mask.sum() < len(df_clean) * 0.1:  # Less than 10% anomalies
                try:
                    window_length = min(11, len(df_clean) // 10 * 2 + 1)
                    if window_length >= 5:
                        polyorder = min(3, window_length // 2)
                        smoothed = savgol_filter(df_clean[col].ffill().bfill(), 
                                                 window_length, polyorder)
                        df_clean.loc[anomaly_mask, col] = smoothed[anomaly_mask]
                except:
                    pass
        
        elif method == 'interpolate':
            # Linear interpolation for gaps
            df_clean.loc[anomaly_mask, col] = np.nan
            df_clean[col] = df_clean[col].interpolate(method='linear', limit_direction='both')
        
        elif method == 'median_filter':
            # Median filter
            window = 5
            for idx in df_clean[anomaly_mask].index:
                start = max(0, idx - window // 2)
                end = min(len(df_clean), idx + window // 2 + 1)
                median_val = df_clean[col].iloc[start:end].median()
                df_clean.loc[idx, col] = median_val
    
    return df_clean


def detect_missing_data_gaps(df: pd.DataFrame, max_gap_seconds: float = 60.0) -> pd.DataFrame:
    """
    Detect missing data gaps in time series.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with 'Time' column
    max_gap_seconds : float
        Maximum allowed gap in seconds before flagging
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with 'Missing_Data_Flag' column
    """
    df_clean = df.copy()
    
    if 'Time' not in df_clean.columns:
        df_clean['Missing_Data_Flag'] = False
        return df_clean
    
    df_clean = df_clean.sort_values('Time').reset_index(drop=True)
    time_diff = df_clean['Time'].diff()
    
    # Flag large gaps
    df_clean['Missing_Data_Flag'] = time_diff > max_gap_seconds
    
    return df_clean


def impute_missing_data(df: pd.DataFrame, method: str = 'linear') -> pd.DataFrame:
    """
    Impute missing data using interpolation.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    method : str
        'linear', 'cubic', or 'forward_fill'
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with imputed values
    """
    df_clean = df.copy()
    
    numeric_cols = ['Voltage', 'Current', 'Capacity', 'Temperature']
    available_cols = [c for c in numeric_cols if c in df_clean.columns]
    
    for col in available_cols:
        if method == 'linear':
            df_clean[col] = df_clean[col].interpolate(method='linear', limit_direction='both')
        elif method == 'cubic' and SCIPY_AVAILABLE:
            df_clean[col] = df_clean[col].interpolate(method='cubic', limit_direction='both')
        elif method == 'forward_fill':
            df_clean[col] = df_clean[col].ffill().bfill()
    
    return df_clean


def verify_electrochemical_anomalies(df: pd.DataFrame) -> pd.DataFrame:
    """
    Verify electrochemical anomalies (Coulombic Efficiency, Capacity Drift).
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with cycle-level or time-series data
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with 'CE_Anomaly' and 'Capacity_Drift_Anomaly' flags
    """
    df_clean = df.copy()
    
    if 'Cycle_Index' not in df_clean.columns:
        df_clean['CE_Anomaly'] = False
        df_clean['Capacity_Drift_Anomaly'] = False
        return df_clean
    
    # Calculate Coulombic Efficiency per cycle
    cycle_data = df_clean.groupby('Cycle_Index').agg({
        'Current': lambda x: x[x > 0].sum() if (x > 0).any() else 0,  # Charge
        'Capacity': ['max', 'min']
    })
    
    if 'Current' in df_clean.columns and 'Capacity' in df_clean.columns:
        # Calculate CE = Discharge Capacity / Charge Capacity
        charge_capacity = df_clean[df_clean['Current'] > 0].groupby('Cycle_Index')['Capacity'].max()
        discharge_capacity = df_clean[df_clean['Current'] < 0].groupby('Cycle_Index')['Capacity'].max()
        
        ce = discharge_capacity / (charge_capacity + 1e-10)
        
        # Flag cycles with CE > 100.5%
        ce_anomaly_cycles = ce[ce > 1.005].index
        
        df_clean['CE_Anomaly'] = df_clean['Cycle_Index'].isin(ce_anomaly_cycles)
        
        # Detect capacity drift/reset
        if len(discharge_capacity) > 1:
            capacity_diff = discharge_capacity.diff()
            # Flag sudden non-monotonic jumps (>20% increase)
            capacity_jump = capacity_diff > (discharge_capacity * 0.2)
            drift_anomaly_cycles = capacity_jump[capacity_jump].index
            
            df_clean['Capacity_Drift_Anomaly'] = df_clean['Cycle_Index'].isin(drift_anomaly_cycles)
        else:
            df_clean['Capacity_Drift_Anomaly'] = False
    else:
        df_clean['CE_Anomaly'] = False
        df_clean['Capacity_Drift_Anomaly'] = False
    
    return df_clean


# ============================================================================
# PHASE 3: Preprocessing for Feature Extraction
# ============================================================================

def detect_time_continuity_issues(df: pd.DataFrame, max_gap_seconds: float = None) -> pd.DataFrame:
    """
    Detect time continuity issues: missing timestamps, duplicates, and large gaps.
    
    ENHANCED v2.0: Preserves original sampling rate and flags issues without modifying data
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with 'Time' column
    max_gap_seconds : float, optional
        Maximum expected gap in seconds. If None, auto-detects from data.
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with 'Time_Continuity_Flag' and 'Gap_Duration' columns
    """
    df_clean = df.copy()
    
    if 'Time' not in df_clean.columns:
        df_clean['Time_Continuity_Flag'] = False
        df_clean['Gap_Duration'] = 0.0
        return df_clean
    
    df_clean = df_clean.sort_values('Time').reset_index(drop=True)
    
    # Calculate time differences
    time_diff = df_clean['Time'].diff()
    
    # Auto-detect expected gap threshold if not provided
    if max_gap_seconds is None:
        # Use median + 3*std as threshold for "normal" gaps
        median_dt = time_diff.median()
        std_dt = time_diff.std()
        max_gap_seconds = median_dt + 3 * std_dt if not pd.isna(std_dt) else median_dt * 2
    
    # Flag large gaps
    df_clean['Gap_Duration'] = time_diff.fillna(0)
    df_clean['Time_Continuity_Flag'] = df_clean['Gap_Duration'] > max_gap_seconds
    
    # Detect duplicate timestamps
    duplicate_times = df_clean['Time'].duplicated(keep=False)
    df_clean.loc[duplicate_times, 'Time_Continuity_Flag'] = True
    
    return df_clean


def analyze_sampling_characteristics(df: pd.DataFrame) -> dict:
    """
    Analyze sampling rate characteristics without modifying data.
    
    ENHANCED v2.0: Provides insights about sampling rate for informed decisions
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with 'Time' column
    
    Returns:
    --------
    dict
        Dictionary with sampling characteristics:
        - sampling_rate_hz: float (average samples per second)
        - is_uniform: bool (whether sampling is uniform)
        - min_interval: float (minimum time interval)
        - max_interval: float (maximum time interval)
        - median_interval: float (median time interval)
        - std_interval: float (standard deviation of intervals)
        - total_samples: int
        - total_duration: float (total time span)
    """
    if 'Time' not in df.columns:
        return {
            'sampling_rate_hz': 0.0,
            'is_uniform': False,
            'min_interval': 0.0,
            'max_interval': 0.0,
            'median_interval': 0.0,
            'std_interval': 0.0,
            'total_samples': len(df),
            'total_duration': 0.0
        }
    
    df_sorted = df.sort_values('Time').reset_index(drop=True)
    time_diff = df_sorted['Time'].diff().dropna()
    
    if len(time_diff) == 0:
        return {
            'sampling_rate_hz': 0.0,
            'is_uniform': False,
            'min_interval': 0.0,
            'max_interval': 0.0,
            'median_interval': 0.0,
            'std_interval': 0.0,
            'total_samples': len(df),
            'total_duration': 0.0
        }
    
    median_interval = float(time_diff.median())
    std_interval = float(time_diff.std())
    min_interval = float(time_diff.min())
    max_interval = float(time_diff.max())
    
    # Determine if sampling is uniform (coefficient of variation < 0.1)
    cv = std_interval / (median_interval + 1e-10)
    is_uniform = cv < 0.1
    
    # Calculate average sampling rate
    sampling_rate_hz = 1.0 / (median_interval + 1e-10) if median_interval > 0 else 0.0
    
    # Total duration
    total_duration = float(df_sorted['Time'].max() - df_sorted['Time'].min())
    
    return {
        'sampling_rate_hz': sampling_rate_hz,
        'is_uniform': is_uniform,
        'min_interval': min_interval,
        'max_interval': max_interval,
        'median_interval': median_interval,
        'std_interval': std_interval,
        'total_samples': len(df),
        'total_duration': total_duration
    }


def intelligent_resample_with_flagging(df: pd.DataFrame, target_frequency_hz: float = None, 
                                       preserve_original: bool = True) -> Tuple[pd.DataFrame, dict]:
    """
    Intelligently resample data with flagging of interpolated points.
    
    ENHANCED v2.0: Only resamples if necessary and flags all interpolated data
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with 'Time' column
    target_frequency_hz : float, optional
        Target sampling frequency. If None, uses original rate.
    preserve_original : bool
        If True, keeps original data and adds resampled columns with suffix '_resampled'
    
    Returns:
    --------
    Tuple[pd.DataFrame, dict]
        Resampled dataframe and metadata dict with:
        - interpolation_percentage: float (percentage of interpolated points)
        - original_samples: int
        - resampled_samples: int
        - original_rate_hz: float
        - target_rate_hz: float
    """
    if 'Time' not in df.columns:
        return df, {'interpolation_percentage': 0.0, 'original_samples': len(df), 
                   'resampled_samples': len(df), 'original_rate_hz': 0.0, 'target_rate_hz': 0.0}
    
    # Analyze original sampling
    sampling_info = analyze_sampling_characteristics(df)
    original_rate_hz = sampling_info['sampling_rate_hz']
    
    # If no target frequency specified, preserve original
    if target_frequency_hz is None:
        target_frequency_hz = original_rate_hz
    
    # If already at target frequency and uniform, no resampling needed
    if (abs(target_frequency_hz - original_rate_hz) < 0.01 and 
        sampling_info['is_uniform'] and 
        not preserve_original):
        return df, {
            'interpolation_percentage': 0.0,
            'original_samples': len(df),
            'resampled_samples': len(df),
            'original_rate_hz': original_rate_hz,
            'target_rate_hz': target_frequency_hz
        }
    
    # Perform resampling
    df_resampled = resample_to_uniform_frequency(df, target_frequency_hz)
    
    # Add interpolation flag
    df_resampled['Is_Interpolated'] = False
    
    # Mark interpolated points (points not in original time grid)
    if 'Time' in df_resampled.columns:
        original_times = set(df['Time'].values)
        interpolated_mask = ~df_resampled['Time'].isin(original_times)
        df_resampled.loc[interpolated_mask, 'Is_Interpolated'] = True
    
    interpolation_percentage = (df_resampled['Is_Interpolated'].sum() / len(df_resampled)) * 100
    
    metadata = {
        'interpolation_percentage': float(interpolation_percentage),
        'original_samples': len(df),
        'resampled_samples': len(df_resampled),
        'original_rate_hz': float(original_rate_hz),
        'target_rate_hz': float(target_frequency_hz)
    }
    
    if preserve_original:
        # Keep original columns and add resampled versions
        for col in ['Current', 'Voltage', 'Capacity', 'Temperature']:
            if col in df_resampled.columns:
                df_resampled[f'{col}_resampled'] = df_resampled[col]
                if col in df.columns:
                    # Reindex original data to resampled time grid using nearest neighbor
                    df_resampled[col] = df[col].reindex(df_resampled.index, method='nearest')
    
    return df_resampled, metadata


def resample_to_uniform_capacity_axis(df: pd.DataFrame, capacity_points: int = 1000) -> pd.DataFrame:
    """
    Resample voltage data to uniform capacity axis.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with Voltage and Capacity columns
    capacity_points : int
        Number of points in uniform capacity grid
    
    Returns:
    --------
    pd.DataFrame
        Resampled dataframe with uniform capacity axis
    """
    if 'Voltage' not in df.columns or 'Capacity' not in df.columns:
        return df
    
    df_clean = df.copy()
    
    # Create uniform capacity grid
    if 'Cycle_Index' in df_clean.columns:
        resampled_dfs = []
        
        for cycle in df_clean['Cycle_Index'].unique():
            cycle_data = df_clean[df_clean['Cycle_Index'] == cycle].copy()
            
            if len(cycle_data) < 2:
                continue
            
            # Sort by capacity
            cycle_data = cycle_data.sort_values('Capacity').reset_index(drop=True)
            
            cap_min = cycle_data['Capacity'].min()
            cap_max = cycle_data['Capacity'].max()
            
            if cap_max <= cap_min:
                continue
            
            # Create uniform capacity grid
            uniform_capacity = np.linspace(cap_min, cap_max, capacity_points)
            
            # Interpolate voltage
            if SCIPY_AVAILABLE:
                f_voltage = interp1d(cycle_data['Capacity'].values, 
                                    cycle_data['Voltage'].values,
                                    kind='linear', 
                                    bounds_error=False, 
                                    fill_value='extrapolate')
                uniform_voltage = f_voltage(uniform_capacity)
            else:
                uniform_voltage = np.interp(uniform_capacity, 
                                          cycle_data['Capacity'].values,
                                          cycle_data['Voltage'].values)
            
            # Create resampled dataframe
            resampled = pd.DataFrame({
                'Cycle_Index': cycle,
                'Capacity': uniform_capacity,
                'Voltage': uniform_voltage
            })
            
            resampled_dfs.append(resampled)
        
        if resampled_dfs:
            return pd.concat(resampled_dfs, ignore_index=True)
    
    return df_clean


def resample_to_uniform_voltage_axis(df: pd.DataFrame, voltage_points: int = 1000) -> pd.DataFrame:
    """
    Resample capacity data to uniform voltage axis.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with Voltage and Capacity columns
    voltage_points : int
        Number of points in uniform voltage grid
    
    Returns:
    --------
    pd.DataFrame
        Resampled dataframe with uniform voltage axis
    """
    if 'Voltage' not in df.columns or 'Capacity' not in df.columns:
        return df
    
    df_clean = df.copy()
    
    if 'Cycle_Index' in df_clean.columns:
        resampled_dfs = []
        
        for cycle in df_clean['Cycle_Index'].unique():
            cycle_data = df_clean[df_clean['Cycle_Index'] == cycle].copy()
            
            if len(cycle_data) < 2:
                continue
            
            # Sort by voltage
            cycle_data = cycle_data.sort_values('Voltage').reset_index(drop=True)
            
            volt_min = cycle_data['Voltage'].min()
            volt_max = cycle_data['Voltage'].max()
            
            if volt_max <= volt_min:
                continue
            
            # Create uniform voltage grid
            uniform_voltage = np.linspace(volt_min, volt_max, voltage_points)
            
            # Interpolate capacity
            if SCIPY_AVAILABLE:
                f_capacity = interp1d(cycle_data['Voltage'].values,
                                     cycle_data['Capacity'].values,
                                     kind='linear',
                                     bounds_error=False,
                                     fill_value='extrapolate')
                uniform_capacity = f_capacity(uniform_voltage)
            else:
                uniform_capacity = np.interp(uniform_voltage,
                                           cycle_data['Voltage'].values,
                                           cycle_data['Capacity'].values)
            
            # Create resampled dataframe
            resampled = pd.DataFrame({
                'Cycle_Index': cycle,
                'Voltage': uniform_voltage,
                'Capacity': uniform_capacity
            })
            
            resampled_dfs.append(resampled)
        
        if resampled_dfs:
            return pd.concat(resampled_dfs, ignore_index=True)
    
    return df_clean


def resample_to_uniform_frequency(df: pd.DataFrame, frequency_hz: float = 1.0) -> pd.DataFrame:
    """
    Resample time-series data to uniform frequency (Issue 3 Fix, v1.2 Bug Fix).
    Essential for ICA/dQ/dV analysis which requires uniform sampling.
    
    FIXES in v1.2:
    - Cycle_Index, Step_Index, Data_Point are NOT interpolated (categorical/discrete)
    - Only continuous physical signals are interpolated (Current, Voltage, Capacity, etc.)
    - Categorical columns use forward-fill or nearest neighbor
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe with Time column
    frequency_hz : float
        Target sampling frequency in Hz (default: 1.0 Hz = 1 sample per second)
    
    Returns:
    --------
    pd.DataFrame
        Resampled dataframe with uniform time spacing
    """
    if 'Time' not in df.columns:
        return df
    
    df_clean = df.copy()
    df_clean = df_clean.sort_values('Time').reset_index(drop=True)
    
    # Columns to EXCLUDE from interpolation (categorical/discrete) - v1.2 Fix
    exclude_from_interpolation = [
        'Cycle_Index', 'Step_Index', 'Data_Point', 'Step_Time', 'Step_Time(s)',
        'Is_RPT', 'Is_FC_Data', 'Anomaly_Flag', 'Missing_Data_Flag',
        'CE_Anomaly', 'Capacity_Drift_Anomaly', 'Step_Type'
    ]
    
    # Also exclude any column with "flag", "index", "point" in name (case-insensitive)
    exclude_patterns = ['flag', 'index', 'point', 'type', 'id', 'source']
    for col in df_clean.columns:
        col_lower = col.lower()
        if any(pattern in col_lower for pattern in exclude_patterns):
            if col not in exclude_from_interpolation:
                exclude_from_interpolation.append(col)
    
    # Columns to interpolate (continuous physical signals)
    numeric_cols = df_clean.select_dtypes(include=[np.number]).columns.tolist()
    
    # Remove Time and excluded columns
    if 'Time' in numeric_cols:
        numeric_cols.remove('Time')
    
    # Only interpolate truly continuous physical signals
    physical_signals = ['Current', 'Voltage', 'Capacity', 'Temperature', 
                       'Internal_Resistance', 'IR', 'Power', 'Energy', 
                       'SOC', 'SOH', 'Charge_Capacity', 'Discharge_Capacity']
    
    cols_to_interpolate = []
    for col in numeric_cols:
        if col not in exclude_from_interpolation:
            # Include if it's a known physical signal or doesn't match exclusion patterns
            col_lower = col.lower()
            if any(signal.lower() in col_lower for signal in physical_signals):
                cols_to_interpolate.append(col)
            elif not any(pattern in col_lower for pattern in exclude_patterns):
                # Safe to interpolate if not explicitly excluded
                cols_to_interpolate.append(col)
    
    # Create uniform time grid
    time_min = df_clean['Time'].min()
    time_max = df_clean['Time'].max()
    dt = 1.0 / frequency_hz  # Time step in seconds
    uniform_time = np.arange(time_min, time_max + dt, dt)
    
    resampled_data = {'Time': uniform_time}
    
    # Interpolate continuous physical signals
    for col in cols_to_interpolate:
        if SCIPY_AVAILABLE:
            try:
                f_interp = interp1d(
                    df_clean['Time'].values,
                    df_clean[col].values,
                    kind='linear',
                    bounds_error=False,
                    fill_value='extrapolate'
                )
                resampled_data[col] = f_interp(uniform_time)
            except:
                resampled_data[col] = np.interp(
                    uniform_time,
                    df_clean['Time'].values,
                    df_clean[col].values
                )
        else:
            resampled_data[col] = np.interp(
                uniform_time,
                df_clean['Time'].values,
                df_clean[col].values
            )
    
    # Handle categorical/discrete columns with forward-fill or nearest neighbor
    # EXCLUDE Cycle_Index, Step_Index, Data_Point from nearest neighbor - they need special handling
    special_categorical_cols = ['Cycle_Index', 'Step_Index', 'Data_Point']
    categorical_cols = [col for col in df_clean.columns 
                       if col not in cols_to_interpolate and col != 'Time' 
                       and col not in special_categorical_cols]
    
    df_resampled = pd.DataFrame(resampled_data)
    
    # For regular categorical columns, use nearest neighbor (forward-fill approach)
    for col in categorical_cols:
        if col in df_clean.columns:
            # Find nearest original sample for each resampled point
            indices = np.searchsorted(df_clean['Time'].values, uniform_time, side='left')
            indices = np.clip(indices, 0, len(df_clean) - 1)
            df_resampled[col] = df_clean[col].iloc[indices].values
    
    # CRITICAL FIX: Set Cycle_Index once per cycle (v1.2 Bug Fix #1)
    # Don't use nearest neighbor - assign based on time boundaries
    if 'Cycle_Index' in df_clean.columns:
        # Get unique cycles from original data, ensure they're integers
        unique_cycles_raw = df_clean['Cycle_Index'].dropna().unique()
        # Convert to integers (handle any fractional values)
        unique_cycles = sorted([int(round(cyc)) for cyc in unique_cycles_raw])
        
        # Initialize Cycle_Index column in resampled data
        df_resampled['Cycle_Index'] = 0  # Initialize with 0
        
        # Map resampled points to original cycles based on time boundaries
        for cycle_id in unique_cycles:
            # Find all original rows with this cycle (handle fractional matches)
            cycle_mask = (df_clean['Cycle_Index'] >= cycle_id - 0.5) & (df_clean['Cycle_Index'] < cycle_id + 0.5)
            if cycle_mask.any():
                cycle_time_min = df_clean.loc[cycle_mask, 'Time'].min()
                cycle_time_max = df_clean.loc[cycle_mask, 'Time'].max()
                # Assign cycle to resampled points in this time range
                cycle_mask_resampled = (df_resampled['Time'] >= cycle_time_min) & (df_resampled['Time'] <= cycle_time_max)
                df_resampled.loc[cycle_mask_resampled, 'Cycle_Index'] = cycle_id
        
        # Ensure Cycle_Index is integer
        df_resampled['Cycle_Index'] = df_resampled['Cycle_Index'].astype(int)
    
    # Handle Step_Index and Data_Point similarly (v1.2 Bug Fix #2)
    # Step_Index: Use forward-fill within each cycle
    if 'Step_Index' in df_clean.columns:
        df_resampled['Step_Index'] = 0  # Initialize
        if 'Cycle_Index' in df_resampled.columns:
            for cycle_id in df_resampled['Cycle_Index'].unique():
                cycle_mask_resampled = df_resampled['Cycle_Index'] == cycle_id
                cycle_mask_original = (df_clean['Cycle_Index'] >= cycle_id - 0.5) & (df_clean['Cycle_Index'] < cycle_id + 0.5)
                if cycle_mask_original.any():
                    # Use nearest neighbor for Step_Index within cycle
                    cycle_times = df_resampled.loc[cycle_mask_resampled, 'Time'].values
                    original_times = df_clean.loc[cycle_mask_original, 'Time'].values
                    original_steps = df_clean.loc[cycle_mask_original, 'Step_Index'].values
                    indices = np.searchsorted(original_times, cycle_times, side='left')
                    indices = np.clip(indices, 0, len(original_times) - 1)
                    df_resampled.loc[cycle_mask_resampled, 'Step_Index'] = original_steps[indices]
            # Ensure Step_Index is integer
            df_resampled['Step_Index'] = df_resampled['Step_Index'].astype(int)
    
    # Data_Point: Not meaningful on resampled grid, but preserve if needed
    if 'Data_Point' in df_clean.columns:
        # Data_Point is just an index - not meaningful after resampling
        # We could drop it or set to NaN, but for compatibility, use nearest neighbor
        indices = np.searchsorted(df_clean['Time'].values, uniform_time, side='left')
        indices = np.clip(indices, 0, len(df_clean) - 1)
        df_resampled['Data_Point'] = df_clean['Data_Point'].iloc[indices].values
    
    # Add per-cycle relative time (t_cycle) - v1.2 Feature #4
    if 'Cycle_Index' in df_resampled.columns:
        df_resampled['t_cycle'] = df_resampled.groupby('Cycle_Index')['Time'].transform(
            lambda s: s - s.min()
        )
    
    return df_resampled


def smooth_for_derivatives(df: pd.DataFrame, window_length: int = 51, polyorder: int = 3, apply_to_original: bool = True) -> pd.DataFrame:
    """
    Apply Savitzky-Golay filter for smoothing and spike removal (Issue 4 Fix).
    By default, applies smoothing directly to Voltage and Current columns to remove noise.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    window_length : int
        Window length for Savitzky-Golay filter (must be odd). Default 51 for better noise removal.
    polyorder : int
        Polynomial order. Default 3.
    apply_to_original : bool
        If True, smooths the original Voltage/Current columns. If False, creates _Smoothed columns.
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with smoothed values
    """
    if not SCIPY_AVAILABLE:
        return df
    
    df_clean = df.copy()
    
    # Ensure window_length is odd
    if window_length % 2 == 0:
        window_length += 1
    
    # Minimum window size check
    min_window = max(5, polyorder + 2)
    if window_length < min_window:
        window_length = min_window
        if window_length % 2 == 0:
            window_length += 1
    
    # Smooth voltage and current to remove spikes (Issue 4)
    if 'Voltage' in df_clean.columns and len(df_clean) >= window_length:
        try:
            voltage_smoothed = savgol_filter(
                df_clean['Voltage'].ffill().bfill().values,
                window_length,
                polyorder
            )
            if apply_to_original:
                df_clean['Voltage'] = voltage_smoothed
            else:
                df_clean['Voltage_Smoothed'] = voltage_smoothed
        except Exception as e:
            if not apply_to_original:
                df_clean['Voltage_Smoothed'] = df_clean['Voltage']
    
    if 'Current' in df_clean.columns and len(df_clean) >= window_length:
        try:
            current_smoothed = savgol_filter(
                df_clean['Current'].ffill().bfill().values,
                window_length,
                polyorder
            )
            if apply_to_original:
                df_clean['Current'] = current_smoothed
            else:
                df_clean['Current_Smoothed'] = current_smoothed
        except Exception as e:
            if not apply_to_original:
                df_clean['Current_Smoothed'] = df_clean['Current']
    
    if 'Capacity' in df_clean.columns and len(df_clean) >= window_length:
        try:
            capacity_smoothed = savgol_filter(
                df_clean['Capacity'].ffill().bfill().values,
                window_length,
                polyorder
            )
            if apply_to_original:
                df_clean['Capacity'] = capacity_smoothed
            else:
                df_clean['Capacity_Smoothed'] = capacity_smoothed
        except Exception as e:
            if not apply_to_original:
                df_clean['Capacity_Smoothed'] = df_clean['Capacity']
    
    return df_clean


# ============================================================================
# Canonical Naming (v1.2 Feature #5)
# ============================================================================

def apply_canonical_naming(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply canonical column naming for universal analytics.
    Standardizes to: I, V, Test_Time, t_cycle, Step_Time, DateTime
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with canonical column names
    """
    df_clean = df.copy()
    rename_map = {}
    
    # Map to canonical names
    cols_l = {col.lower(): col for col in df_clean.columns}
    
    # Current -> I
    if 'current' in cols_l and 'I' not in df_clean.columns:
        rename_map[cols_l['current']] = 'I'
    
    # Voltage -> V
    if 'voltage' in cols_l and 'V' not in df_clean.columns:
        rename_map[cols_l['voltage']] = 'V'
    
    # Time -> Test_Time (if t_cycle exists, rename Time to Test_Time)
    if 'time' in cols_l and 'test_time' not in [c.lower() for c in df_clean.columns]:
        if 't_cycle' in df_clean.columns:
            rename_map[cols_l['time']] = 'Test_Time'
        else:
            # If no t_cycle, rename Time to Test_Time
            rename_map[cols_l['time']] = 'Test_Time'
    
    # Step_Time(s) -> Step_Time
    for col in df_clean.columns:
        if col.lower() in ['step_time(s)', 'step_time_s']:
            rename_map[col] = 'Step_Time'
    
    # Apply renames
    if rename_map:
        df_clean = df_clean.rename(columns=rename_map)
    
    return df_clean


# ============================================================================
# Metadata Extraction (Issue 5 Fix)
# ============================================================================

def extract_excel_metadata(file_path_or_buffer, sheet_name: str = None) -> dict:
    """
    Extract metadata from Excel file headers (Issue 5 Fix).
    Extracts test report name, channel number, schedule file, operator, notes.
    
    Parameters:
    -----------
    file_path_or_buffer
        Excel file path or file-like object
    sheet_name : str, optional
        Specific sheet name to extract metadata from
    
    Returns:
    --------
    dict
        Dictionary with extracted metadata
    """
    metadata = {
        'test_report': None,
        'channel': None,
        'schedule_file': None,
        'operator': None,
        'notes': None,
        'format': None,
        'cycler': None
    }
    
    try:
        import openpyxl
        
        if hasattr(file_path_or_buffer, 'read'):
            # File-like object
            wb = openpyxl.load_workbook(file_path_or_buffer, read_only=True, data_only=True)
        else:
            # File path
            wb = openpyxl.load_workbook(file_path_or_buffer, read_only=True, data_only=True)
        
        # Use specified sheet or first sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Read first few rows to find metadata
        for row_idx in range(1, min(20, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            row_text = ' '.join([str(v) for v in row_values if v is not None]).lower()
            
            # Look for metadata patterns
            for i, cell in enumerate(ws[row_idx]):
                if cell.value is None:
                    continue
                
                cell_text = str(cell.value).lower()
                
                # Test report
                if 'test report' in cell_text or 'report name' in cell_text:
                    if i + 1 < len(row_values) and row_values[i + 1]:
                        metadata['test_report'] = str(row_values[i + 1])
                
                # Channel
                if 'channel' in cell_text:
                    if i + 1 < len(row_values) and row_values[i + 1]:
                        metadata['channel'] = str(row_values[i + 1])
                
                # Schedule file
                if 'schedule' in cell_text and 'file' in cell_text:
                    if i + 1 < len(row_values) and row_values[i + 1]:
                        metadata['schedule_file'] = str(row_values[i + 1])
                
                # Operator
                if 'operator' in cell_text:
                    if i + 1 < len(row_values) and row_values[i + 1]:
                        metadata['operator'] = str(row_values[i + 1])
                
                # Notes
                if 'note' in cell_text or 'comment' in cell_text:
                    if i + 1 < len(row_values) and row_values[i + 1]:
                        metadata['notes'] = str(row_values[i + 1])
                
                # Format/Cycler detection
                if 'calce' in cell_text:
                    metadata['format'] = 'calce'
                    metadata['cycler'] = 'CALCE'
                elif 'arbin' in cell_text:
                    metadata['format'] = 'arbin'
                    metadata['cycler'] = 'Arbin'
                elif 'biologic' in cell_text:
                    metadata['format'] = 'biologic'
                    metadata['cycler'] = 'Biologic'
        
        wb.close()
        
    except Exception as e:
        # If extraction fails, return empty metadata
        pass
    
    return metadata


# ============================================================================
# Main Cleaning Pipeline
# ============================================================================

def clean_dataframe(df: pd.DataFrame, options: dict, metadata: dict = None) -> pd.DataFrame:
    """
    Comprehensive battery data cleaning pipeline implementing 3-phase approach.
    Universal Cleaning Module v1.1 - Handles CALCE, Severson .mat, and generic formats.
    
    Phase 1: Data Ingestion and Harmonization
    Phase 2: Anomaly Detection and Correction
    Phase 3: Preprocessing for Feature Extraction
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    options : dict
        Cleaning options:
        - harmonize_schema: bool (default: True)
        - detect_cycler: bool (default: True)
        - segment_steps: bool (default: True)
        - assign_step_types: bool (default: True)
        - identify_rpt: bool (default: True)
        - detect_anomalies: bool (default: True)
        - anomaly_method: str ('zscore' or 'lof', default: 'zscore')
        - correct_anomalies: bool (default: True)
        - correction_method: str ('savgol', 'interpolate', 'median_filter', default: 'savgol')
        - detect_missing_data: bool (default: True)
        - impute_missing: bool (default: True)
        - imputation_method: str ('linear', 'cubic', 'forward_fill', default: 'linear')
        - verify_electrochemical: bool (default: True)
        - resample_uniform_frequency: bool (default: False) - ENHANCED v2.0: Preserves original sampling by default
        - frequency_hz: float (default: None) - Target sampling frequency. None preserves original rate
        - preserve_original_sampling: bool (default: True) - Keep original data when resampling
        - detect_time_continuity: bool (default: True) - Detect time gaps and duplicates
        - resample_capacity_axis: bool (default: False)
        - capacity_points: int (default: 1000)
        - resample_voltage_axis: bool (default: False)
        - voltage_points: int (default: 1000)
        - smooth_for_derivatives: bool (default: True) - NEW: Default enabled
        - savgol_window: int (default: 51) - NEW: Larger default for better smoothing
        - savgol_polyorder: int (default: 3)
    metadata : dict, optional
        Metadata dictionary from Excel headers (test_report, channel, schedule_file, etc.)
    
    Returns:
    --------
    pd.DataFrame
        Cleaned dataframe
    """
    df_clean = df.copy()
    
    # Phase 1: Data Ingestion and Harmonization
    if options.get('harmonize_schema', True):
        detect_cycler = options.get('detect_cycler', True)
        cycler_format = detect_cycler_format(df_clean, metadata) if detect_cycler else None
        df_clean = harmonize_to_standard_schema(df_clean, cycler_format)
    
    # Phase 2: Anomaly Detection and Correction
    if options.get('segment_steps', True):
        df_clean = segment_steps_and_cycles(df_clean)
    
    if options.get('assign_step_types', True):
        df_clean = assign_step_types(df_clean)
    
    if options.get('identify_rpt', True):
        df_clean = identify_rpt_cycles(df_clean)
    
    if options.get('detect_anomalies', True):
        window_size = options.get('anomaly_window_size', 10)
        method = options.get('anomaly_method', 'zscore')
        df_clean = detect_statistical_anomalies(df_clean, window_size, method)
    
    if options.get('correct_anomalies', True):
        correction_method = options.get('correction_method', 'savgol')
        df_clean = correct_anomalies(df_clean, correction_method)
    
    if options.get('detect_missing_data', True):
        max_gap = options.get('max_gap_seconds', 60.0)
        df_clean = detect_missing_data_gaps(df_clean, max_gap)
    
    if options.get('impute_missing', True):
        imputation_method = options.get('imputation_method', 'linear')
        df_clean = impute_missing_data(df_clean, imputation_method)
    
    if options.get('verify_electrochemical', True):
        df_clean = verify_electrochemical_anomalies(df_clean)
    
    # Phase 3: Preprocessing for Feature Extraction
    if options.get('resample_capacity_axis', False):
        capacity_points = options.get('capacity_points', 1000)
        df_clean = resample_to_uniform_capacity_axis(df_clean, capacity_points)
    
    if options.get('resample_voltage_axis', False):
        voltage_points = options.get('voltage_points', 1000)
        df_clean = resample_to_uniform_voltage_axis(df_clean, voltage_points)
    
    # Apply default Savitzky-Golay smoothing for spike removal (Issue 4 Fix)
    # This is now enabled by default for better noise removal
    if options.get('smooth_for_derivatives', True):  # Default True now
        window_length = options.get('savgol_window', 51)  # Larger default window
        polyorder = options.get('savgol_polyorder', 3)
        apply_to_original = options.get('savgol_apply_to_original', True)
        df_clean = smooth_for_derivatives(df_clean, window_length, polyorder, apply_to_original)
    
    # Detect time continuity issues (ENHANCED v2.0: Preserve original sampling)
    if options.get('detect_time_continuity', True):
        df_clean = detect_time_continuity_issues(df_clean)
    
    # Analyze sampling characteristics (ENHANCED v2.0: Inform user about sampling rate)
    sampling_info = analyze_sampling_characteristics(df_clean)
    if metadata is None:
        metadata = {}
    metadata['sampling_characteristics'] = sampling_info
    
    # Resample to uniform frequency (ENHANCED v2.0: Optional, preserves original by default)
    # Only resample if explicitly requested or if data is highly irregular
    should_resample = options.get('resample_uniform_frequency', False)  # Default False now
    if not should_resample and not sampling_info['is_uniform']:
        # Auto-resample only if sampling is very irregular (CV > 0.5)
        cv = sampling_info['std_interval'] / (sampling_info['median_interval'] + 1e-10)
        should_resample = cv > 0.5
    
    if should_resample:
        frequency_hz = options.get('frequency_hz', None)  # None = preserve original rate
        preserve_original = options.get('preserve_original_sampling', True)  # Default True
        
        if frequency_hz is None:
            # Use original sampling rate
            frequency_hz = sampling_info['sampling_rate_hz']
        
        df_clean, resample_metadata = intelligent_resample_with_flagging(
            df_clean, target_frequency_hz=frequency_hz, preserve_original=preserve_original
        )
        
        # Store resampling metadata
        metadata['resampling'] = resample_metadata
        if resample_metadata['interpolation_percentage'] > 0:
            metadata['interpolation_warning'] = (
                f"{resample_metadata['interpolation_percentage']:.1f}% of data points were interpolated"
            )
    
    # Apply range filters (replace out-of-range values with 0)
    df_clean = apply_range_filters(df_clean, options)
    
    # v1.2 Feature #5: Canonical naming for universal analytics
    df_clean = apply_canonical_naming(df_clean)
    
    # Preserve metadata as attributes (Issue 5 Fix)
    if metadata:
        for key, value in metadata.items():
            if value is not None:
                df_clean.attrs[f'metadata_{key}'] = value
    
    return df_clean


# ============================================================================
# Utility Functions
# ============================================================================

def get_dataframe_info(df: pd.DataFrame) -> dict:
    """Get summary information about a dataframe."""
    info = {
        'shape': df.shape,
        'columns': list(df.columns),
        'dtypes': df.dtypes.to_dict(),
        'null_counts': df.isnull().sum().to_dict(),
        'null_percentages': (df.isnull().sum() / len(df) * 100).to_dict(),
        'duplicate_rows': df.duplicated().sum(),
        'memory_usage': df.memory_usage(deep=True).sum() / 1024 / 1024
    }
    return info


def to_csv_download(df: pd.DataFrame, filename: str = "cleaned_data.csv", remove_sheet_source: bool = False) -> bytes:
    """Convert dataframe to CSV bytes for download."""
    df_export = df.copy()
    
    if remove_sheet_source and 'Sheet_Source' in df_export.columns:
        df_export = df_export.drop(columns=['Sheet_Source'])
    
    output = BytesIO()
    df_export.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    return output.getvalue()


def validate_analysis_compatibility(df: pd.DataFrame) -> dict:
    """Validate if cleaned dataframe is compatible with analysis module."""
    cols_l = [c.lower() for c in df.columns]
    warnings = []
    missing_columns = []
    
    has_cycle = any(c in ["cycle", "cycle_number", "cycle_idx", "cycle_index"] for c in cols_l)
    has_discharge_cap = any(
        "discharge_capacity" in c or "discharge_cap" in c or "qdischarge" in c or
        (c in ["capacity", "cap", "q"] and "discharge" in " ".join(cols_l)) or c == "qdischarge"
        for c in cols_l)
    
    has_voltage = any(c in ["v", "volt", "voltage", "voltage_v"] for c in cols_l)
    has_capacity = any("capacity" in c or c in ["q", "ah", "mah", "capacity_ah", "capacity_mah"] for c in cols_l)
    
    format_type = 'unknown'
    is_compatible = False
    
    if has_cycle and has_discharge_cap:
        format_type = 'cycle_level'
        is_compatible = True
    elif has_voltage and has_capacity:
        format_type = 'voltage_capacity'
        is_compatible = True
    
    if 'Sheet_Source' in df.columns:
        warnings.append("'Sheet_Source' column detected. Consider removing it for analysis compatibility.")
    
    return {
        'is_compatible': is_compatible,
        'format_type': format_type,
        'warnings': warnings,
        'missing_columns': missing_columns
    }


def get_excel_sheets(file_path_or_buffer) -> list:
    """Get list of sheet names from an Excel file."""
    try:
        excel_file = pd.ExcelFile(file_path_or_buffer)
        return excel_file.sheet_names
    except Exception:
        return []


def read_excel_sheets(file_path_or_buffer, sheet_names: list = None) -> dict:
    """Read specific sheets from an Excel file."""
    try:
        excel_file = pd.ExcelFile(file_path_or_buffer)
        if sheet_names is None:
            sheet_names = excel_file.sheet_names
        
        sheets_dict = {}
        for sheet_name in sheet_names:
            if sheet_name in excel_file.sheet_names:
                sheets_dict[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
        return sheets_dict
    except Exception:
        return {}


def filter_channel_sheets(sheet_names: list) -> list:
    """Filter sheet names to get only Channel_ sheets (exclude 'info')."""
    channel_sheets = []
    for sheet in sheet_names:
        if sheet.lower() != 'info' and sheet.lower().startswith('channel_'):
            channel_sheets.append(sheet)
    return sorted(channel_sheets, key=lambda x: int(x.split('_')[-1]) if x.split('_')[-1].isdigit() else 999)


def apply_range_filters(df: pd.DataFrame, options: dict) -> pd.DataFrame:
    """Apply user-specified range filters - replace out-of-range values with 0."""
    if not options.get('apply_range_filters', False):
        return df
    
    df_clean = df.copy()
    range_filters = options.get('range_filters', {})
    
    for col_name, range_config in range_filters.items():
        if col_name in df_clean.columns and pd.api.types.is_numeric_dtype(df_clean[col_name]):
            min_val = range_config.get('min', None)
            max_val = range_config.get('max', None)
            replace_with = range_config.get('replace_with', 0.0)
            
            mask = pd.Series(True, index=df_clean.index)
            
            if min_val is not None:
                mask = mask & (df_clean[col_name] >= min_val)
            if max_val is not None:
                mask = mask & (df_clean[col_name] <= max_val)
            
            df_clean.loc[~mask, col_name] = replace_with
    
    return df_clean


def get_battery_default_ranges() -> dict:
    """Get default value ranges for common battery parameters."""
    return {
        'voltage': {'min': 2.0, 'max': 4.5, 'replace_with': 0.0, 'unit': 'V'},
        'current': {'min': -50.0, 'max': 50.0, 'replace_with': 0.0, 'unit': 'A'},
        'capacity': {'min': 0.0, 'max': 10.0, 'replace_with': 0.0, 'unit': 'Ah'},
        'discharge_capacity': {'min': 0.0, 'max': 10.0, 'replace_with': 0.0, 'unit': 'Ah'},
        'charge_capacity': {'min': 0.0, 'max': 10.0, 'replace_with': 0.0, 'unit': 'Ah'},
        'temperature': {'min': -40.0, 'max': 90.0, 'replace_with': 25.0, 'unit': '°C'},
        'internal_resistance': {'min': 0.0, 'max': 1.0, 'replace_with': 0.0, 'unit': 'Ohm'},
        'soc': {'min': 0.0, 'max': 100.0, 'replace_with': 0.0, 'unit': '%'},
        'soh': {'min': 0.0, 'max': 100.0, 'replace_with': 100.0, 'unit': '%'},
        'cycle': {'min': 1, 'max': 10000, 'replace_with': 0, 'unit': ''},
        'power': {'min': -500.0, 'max': 500.0, 'replace_with': 0.0, 'unit': 'W'},
        'energy': {'min': 0.0, 'max': 1000.0, 'replace_with': 0.0, 'unit': 'Wh'},
    }


def export_to_cell11_mat_format(df: pd.DataFrame, cell_id: str = "cell11") -> dict:
    """Convert dataframe to cell11dataset_for_python.mat format structure."""
    mat_data = {}
    
    cols_l = [c.lower() for c in df.columns]
    
    cycle_col = None
    for i, c in enumerate(cols_l):
        if c in ["cycle", "cycle_number", "cycle_idx", "cycle_index"]:
            cycle_col = df.columns[i]
            break
    
    cap_col = None
    for i, c in enumerate(cols_l):
        if "discharge_capacity" in c or "discharge_cap" in c or "qdischarge" in c:
            cap_col = df.columns[i]
            break
    if cap_col is None:
        for i, c in enumerate(cols_l):
            if c in ["capacity", "cap", "q"]:
                cap_col = df.columns[i]
                break
    
    voltage_col = None
    current_col = None
    for i, c in enumerate(cols_l):
        if ('voltage' in c or c in ['v', 'volt']) and voltage_col is None:
            voltage_col = df.columns[i]
        if ('current' in c or c in ['a', 'amps', 'i']) and current_col is None:
            current_col = df.columns[i]
    
    if cycle_col and cap_col:
        summary_cols = [cycle_col, cap_col]
        
        for col in df.columns:
            col_l = col.lower()
            if any(x in col_l for x in ['internal_resistance', 'ir', 'resistance', 'temperature', 'temp', 'soc', 'soh']):
                if col not in summary_cols:
                    summary_cols.append(col)
        
        summary = df[summary_cols].copy()
        summary.columns = [c.lower().replace(' ', '_') for c in summary.columns]
        mat_data['summary'] = summary.values if len(summary) > 0 else np.array([])
    
    if cycle_col:
        cycles = {}
        cycles_time_series = {}
        
        unique_cycles = df[cycle_col].dropna().unique()
        
        for cycle_num in unique_cycles:
            cycle_data = df[df[cycle_col] == cycle_num]
            
            cycle_dict = {}
            for col in df.columns:
                if col != cycle_col:
                    values = cycle_data[col].dropna().values
                    if len(values) > 0:
                        cycle_dict[col] = values[0] if len(values) == 1 else values
            
            if cycle_dict:
                cycles[int(cycle_num)] = cycle_dict
            
            if voltage_col and current_col:
                ts_data = {}
                if voltage_col in cycle_data.columns:
                    ts_data['voltage'] = cycle_data[voltage_col].dropna().values
                if current_col in cycle_data.columns:
                    ts_data['current'] = cycle_data[current_col].dropna().values
                
                if ts_data:
                    cycles_time_series[int(cycle_num)] = ts_data
        
        if cycles:
            mat_data['cycles'] = cycles
        if cycles_time_series:
            mat_data['cycles_time_series'] = cycles_time_series
    
    mat_data['policy'] = f"{cell_id}_policy"
    mat_data['policy_readable'] = f"Cleaned data from {cell_id}"
    if cycle_col and len(df) > 0:
        mat_data['cycle_life'] = int(df[cycle_col].max()) if cycle_col in df.columns else 0
    
    return mat_data


def to_mat_download(df: pd.DataFrame, filename: str = "cleaned_data.mat", cell_id: str = "cell11") -> bytes:
    """Convert dataframe to .mat file bytes in cell11dataset_for_python format."""
    try:
        from scipy.io import savemat
        import io
        
        mat_data = export_to_cell11_mat_format(df, cell_id)
        
        output = io.BytesIO()
        savemat(output, mat_data, format='5', oned_as='column')
        output.seek(0)
        return output.getvalue()
    except ImportError:
        raise ImportError("SciPy is required for .mat export. Install with: pip install scipy")
